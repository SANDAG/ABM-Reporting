import openpyxl
import pandas as pd
import pyodbc
import settings  # import python settings file
import time


# user inputs a list of ABM database scenario_id
input_string = input(("Enter a list of ABM [scenario_id]s separated by "
                      "spaces (e.g. 1 2 3)"))
scenarios = input_string.split()

# initialize Federal RTP 2020 Performance Measure Excel workbook template
template = openpyxl.load_workbook(settings.template_path)

# build SQL Server connection string
conn = pyodbc.connect("DRIVER={SQL Server Native Client 11.0};SERVER="
                      + settings.server +
                      ";DATABASE=" +
                      settings.database +
                      ";Trusted_Connection=yes;")


# start execution timer
start = time.time()

# for each [scenario_id] -----------------------------------------------------
for i, scenario_id in enumerate(scenarios):
    print("Beginning [scenario_id]: " + scenario_id)
    scenario_id = int(scenario_id)
    col = i + 4  # set column in the Excel template to write results

    print("Running: Single-rowset ad-hoc queries")

    # for each dictionary of single-rowset ad-hoc queries --------------------
    for qry in settings.adhoc_queries:

        # execute ad-hoc query
        result = pd.read_sql_query(
            sql=qry["query"],
            con=conn,
            params=[scenario_id]
        )

        # for each specified column in the query single-row result set
        # write results to specified Excel template sheet and row
        for column, sheet, row in list(zip(qry["columns"], qry["sheets"], qry["rows"])):
            template[sheet].cell(row=row, column=col).value = \
                result[column][0]

    print("Elapsed time: %5.2f minutes" % ((time.time() - start) / 60.0))

    # for each dictionary of Performance Measure stored procedures -----------
    for pm in settings.fed_rtp_20_pm:
        print("Running: " + pm["sp"] + " " + pm["args"])

        # execute stored procedure with specified arguments
        result = pd.read_sql_query(
            sql="EXECUTE " + pm["sp"] + " " + pm["args"],
            con=conn,
            params=[scenario_id]
        )

        # for each specified metric write results
        # to specified Excel template sheet and row
        for metric, sheet, row in list(zip(pm["metrics"], pm["sheets"], pm["rows"])):
            template[sheet].cell(row=row, column=col).value = \
                result[result.metric.str.strip() == metric]["value"].values[0]

        print("Elapsed time: %5.2f minutes" % ((time.time() - start) / 60.0))

    print("Finished [scenario_id]: " + str(scenario_id))

# if protection option set
if settings.template_protect is True:
    # for each worksheet in the template
    for sheet in template.worksheets:
        # set protection and password
        sheet.protection.set_password(settings.template_password)

# save the completed template
template.save(settings.template_write)
