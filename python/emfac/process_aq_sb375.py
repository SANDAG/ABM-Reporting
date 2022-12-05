#from xlrd import open_workbook
from datetime import datetime
import pandas as pd
import openpyxl  # necessary for pandas ExcelWriter
import pyodbc
import csv
import os
import sys
import time
import yaml

#for SB375
#SB375 CO2 
#SB375 VMT
####for SB375 Emission Factor
#SB375 CO2 RunEx
#SB375 CO2 StartEx
#SB 375 vehicle trips

# Start Script Timer
start = time.time()

def find(name, path):
	for root, dirs, files in os.walk(path):
		for file in files:
			if name in file:
				return os.path.join(root, file)

usage = ("Correct Usage: process_aq_sb375.py <EMFAC version: 2014 | 2017> "
        "<Output Folder> <scenario_id1> <scenario_id2> <scenario_id3> ...")
# user inputs
# raise error if emfac version argument passed incorrectly
emfac_version = sys.argv[1]
if emfac_version not in ["2014", "2017"]:
    print(usage)
    sys.exit(-1)

output_folder = sys.argv[2]	
				
# get SQL server attributes from local file
with open("database-specs.yaml", "r", encoding="utf8") as stream:
    try:
        db_info = yaml.safe_load(stream)
    except yaml.YAMLError as exc:
        print(exc)

# set sql server connection string
# noinspection PyArgumentList
sql_con = pyodbc.connect(driver='{SQL Server}',
                         server=db_info['server1']['name'],
                         database=db_info['server1']['db1_1'],
                         trusted_connection='yes')

#base_folder = 'T:\\rtp\\2019rp\\rp19_scen\\abm_runs_bod\\'
#output_folder = '\\output'
#folder = 'T:\\RTP\\2021RP\\2021rp_draft_v1\\abm_runs\\emfac\\emfac2014\\'

with open(output_folder+'/summary_emfac'+ str(emfac_version)+'_SB375_' + str(datetime.now().strftime("%Y-%m-%d")+'.csv'), 'w', newline='') as csvfile:
	writer = csv.writer(csvfile)
	writer.writerow(['Scenario','Scenario_ID','SB375_CO2_runex','SB375_VMT','SB375_CO2_startEx','SB375_vehicle_trip','SB375_CO2'])
					 #scenario,id, sb375_VMT, sb375_CO2,SB375_co2_runex,SB375_co2_runEx,SB375_vehicle_trip])
	for id_argv in sys.argv[3:]:
		scenario_id = int(id_argv)
		scenario = pd.read_sql_query(
			sql=("SELECT RTRIM([name]) AS [name], [year], RTRIM(path) as path  "
			 "FROM [dimension].[scenario] WHERE [scenario_id] = ?"),
			con=sql_con,
			params=[scenario_id]
		)
		#folder = scenario.at[0, "path"] + '\\output'
				
		file_name_pattern = 'EMFAC'+ str(emfac_version) +'-SANDAG-' + scenario.at[0, "name"] + '-' + str(scenario_id) + '-Annual-' + str(scenario.at[0, "year"]) + '-sb375_sb375'
		xlsx = find(file_name_pattern, output_folder)

		if xlsx: 
			wb = openpyxl.load_workbook(xlsx)
			sheet = wb['Total SANDAG']
			sb375_vmt = sheet.cell(2,11).value
			sb375_co2 = sheet.cell(2,16).value
			sb375_vehicle_trip = sheet.cell(2,12).value        
			sb375_co2_runex = sheet.cell(2,13).value
			sb375_co2_startex = sheet.cell(2,15).value
			
			writer.writerow([scenario.at[0, "name"], scenario_id, \
							sb375_co2_runex,sb375_vmt, sb375_co2_startex,sb375_vehicle_trip,sb375_co2])
		else:
			print(f"\nFAILED TO OPEN: {file_name_pattern}.\n If this is unexpected behavior, please re-verify if file exists.\n")

	print(f"\n FILE CREATED AT: {csvfile.name}\n")