# Author: Clint Daniels
# Date: Aug 19, 2015
# To get the key emission and VMT from EMFAC output files for 2015 Regional Plan
# Author: Ziying Ouyang
# Date: Sept 2018
# modified to paramertize EMFAC version and write addtional pollutants by season
# and SB 375 emission related information for 2019 Regional Plan
# Author: Neeco Beltran
# Modified to 

from openpyxl import Workbook
from datetime import datetime
import pandas as pd
import openpyxl  # necessary for pandas ExcelWriter
import pyodbc
import csv
import os
import sys
import time
import yaml

#Annual_VMT
#Annual_trips
#Annual_CO2_emissions
#Summer_ROG
#Summer_NOx
#Annual_SOx
#Winter_CO
#Winter_PM10
#Winter_PM2_5
#Annual_gas_consumption
#Annual_diesel_consumption
#for SB375
#SB375 CO2 
#SB375 VMT
####for SB375 Emission Factor
#SB375 CO2 RunEx
#SB375 CO2 StartEx
#SB 375 vehicle trips

# Start Script Timer
start = time.time()

def find(name, path):
    for root, dirs, files in os.walk(path):
        for file in files:
            if name in file:
                return os.path.join(root, file)

usage = ("Correct Usage: process_aq.py <EMFAC version: 2014 | 2017> "
        "<Output Folder> <scenario_id1> <scenario_id2> <scenario_id3> ...")
# user inputs
# raise error if emfac version argument passed incorrectly
emfac_version = sys.argv[1]
if emfac_version not in ["2014", "2017"]:
    print(usage)
    sys.exit(-1)

output_folder = sys.argv[2] 

# get SQL server attributes from local file
with open("../../resources/emfac/database-specs.yaml", "r", encoding="utf8") as stream:
    try:
        db_info = yaml.safe_load(stream)
    except yaml.YAMLError as exc:
        print(exc)

# set sql server connection string
# noinspection PyArgumentList
sql_con = pyodbc.connect(driver='{SQL Server}',
                         server=db_info['server1']['name'],
                         database=db_info['server1']['db1_1'],
                         trusted_connection='yes')

# folder = 'T:\\RTP\\2021RP\\2021rp_final\\abm_runs\\emfac\\emfac2017\\outputs\\'
folder = output_folder

with open(output_folder+'/summary_emfac'+ str(emfac_version)+'_' + str(datetime.now().strftime("%Y-%m-%d")+'.csv'), 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    #writer.writerow(['Scenario','Scenario_ID','Vehicle Trips','VMT','ROG Budget','Summer ROG Total', 'Summer Adjusted ROG Total', 'NOx Budget', 'Summer NOx Total', 'Summer Adjusted NOx Total'])
                     #scenario,id, annual_vmt, annual_trips,annual_co2, annual_sox, summer_rog, summer_nox, winter_co, winter_pm10, winter_pm2_5, annual_gas, annual_diesel, sb375_VMT, sb375_CO2,SB375_co2_runex,SB375_co2_runEx,SB375_vehicle_trip])
    #writer.writerow(['Scenario','Scenario_id', 'Annual Trips', 'Annual VMT', 'Annual_CO2','Annual PM2.5 Total','Annual PM 10 Total','Annual_gas_consumption','Annual_diesel_consumption','Summer_ROG','Summer_NOx', 'Annual_Gas_VMT', 'Annual_Diesel_VMT'])
    writer.writerow(['Scenario','Scenario_id', 'Annual Trips', 'Annual VMT', 'Annual_CO2','Annual PM2.5 Total','Annual PM 10 Total','Annual_gas_consumption','Annual_diesel_consumption','Summer_ROG','Summer_NOx', 'Winter_CO', 'Annual_Gas_VMT', 'Annual_Diesel_VMT'])
    for id_argv in sys.argv[3:]:
        scenario_id = int(id_argv)
        scenario = pd.read_sql_query(
            sql=("SELECT RTRIM([name]) AS [name], [year], RTRIM(path) as path  "
             "FROM [dimension].[scenario] WHERE [scenario_id] = ?"),
            con=sql_con,
            params=[scenario_id]
        )

        #folder = scenario.at[0, "path"] + '\\output'
        
        file_name_pattern = f'EMFAC{emfac_version}-SANDAG-{scenario.at[0, "name"]}-{str(scenario_id)}-Annual-{str(scenario.at[0, "year"])}_planning'

        xlsx = find(file_name_pattern, folder)
        wb = openpyxl.load_workbook(xlsx)
        sheet = wb['Total SANDAG']
        
        ### Start additions by NBE 10/19/2021
        #### If the Vehicle Type column has the string "GAS", sum the VMT for those values.
        annual_gas_rows = []
        for cell in sheet['H']: 
            if(cell.value is not None):
                if 'GAS' in cell.value:
                    annual_gas_rows.append(cell.row)
        
        annual_gas_values = []
        for i in annual_gas_rows:
            annual_gas_values.append(sheet.cell(i, 11).value)
            
        annual_gas_vmt_value = sum(annual_gas_values)
        
        #### If the Vehicle Type column has the string "DSL", sum the VMT for those values.
        annual_diesel_rows = []
        for cell in sheet['H']:
            if(cell.value is not None):
                if 'DSL' in cell.value:
                    annual_diesel_rows.append(cell.row)
                    
        annual_diesel_values = []
        for i in annual_diesel_rows:
            annual_diesel_values.append(sheet.cell(i, 11).value)
            
        annual_diesel_vmt_value = sum(annual_diesel_values)
        ### End additions by NBE 10/19/2021

        annual_vmt = sheet.cell(2,11).value
        annual_trips = sheet.cell(2,12).value
        
        annual_co2 = sheet.cell(2,42).value
        # annual_sox = sheet.cell(2,60).value
        annual_pm10 = sheet.cell(2,49).value
        annual_pm25 = sheet.cell(2,56).value
        
        annual_gas = sheet.cell(2,61).value
        annual_diesel = sheet.cell(2,62).value

        file_name_pattern = f'EMFAC{emfac_version}-SANDAG-{scenario.at[0, "name"]}-{str(scenario_id)}-Winter-{str(scenario.at[0, "year"])}_planning'
        winter_co = None
        xlsx = find(file_name_pattern, folder)
        if xlsx:
            wb = openpyxl.load_workbook(xlsx)#open_workbook(xlsx)
            sheet = wb['Total SANDAG']
            winter_co = round(sheet.cell(2,34).value,2)
            # winter_pm10 = sheet.cell(1,48).value
            # winter_pm2_5 = sheet.cell(2,56).value
        elif xlsx is None:
            print(f"Failed to open: {file_name_pattern}.\n If this is unexpected behavior, please re-verify if file exists.\n")
        
        file_name_pattern = f'EMFAC{emfac_version}-SANDAG-{scenario.at[0, "name"]}-{str(scenario_id)}-Summer-{str(scenario.at[0, "year"])}_planning'
        summer_rog = None
        summer_nox = None
        xlsx = find(file_name_pattern, folder)
        if xlsx:
            wb = openpyxl.load_workbook(xlsx)
            sheet = wb['Total SANDAG']           
            summer_rog = round(sheet.cell(2,30).value,2)
            summer_nox = round(sheet.cell(2,38).value,2)
            #summer_vmt = sheet.cell(2,11).value
            #summer_trips = sheet.cell(2,12).value
        elif xlsx is None:
            print(f"Failed to open: {file_name_pattern}.\n If this is unexpected behavior, please re-verify if file exists.\n")


        writer.writerow([scenario.at[0, "name"],scenario_id, annual_trips ,annual_vmt, annual_co2, round(annual_pm25,2), round(annual_pm10,2), annual_gas,annual_diesel, (summer_rog), (summer_nox), (winter_co), round(annual_gas_vmt_value), round(annual_diesel_vmt_value)])