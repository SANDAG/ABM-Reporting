import openpyxl
import pandas as pd
import pyodbc
import settings  # import python settings file
import time


# user inputs an ABM database scenario_id and project info
scenario_id = input("Enter an ABM [scenario_id]: ")
proj_name = input("Enter project name: ") # NAVWAR Redevelopment
scenario_header = input("Enter  Scenario Header: ") # Low Density with Transit Center and OT Interchange
geo = input("Enter Report Geography: ") # Project Site
# user inputs filename suffix
file_suffix = input("Enter report file name suffix: ")

# load workbook template
template = openpyxl.load_workbook(settings.template_path)

# initialize output Report based on template
templateWriter = pd.ExcelWriter(
    path=settings.template_path,
    mode="w",
    engine="openpyxl")

templateWriter.book = template
templateWriter.sheets = dict((ws.title, ws) for ws in template.worksheets)

input_ws= template.get_sheet_by_name('Input')
input_ws['C7']= scenario_id
input_ws['C5']= proj_name
input_ws['C8']= scenario_header
input_ws['C9']= geo

# build SQL Server connection string
conn = pyodbc.connect("DRIVER={SQL Server Native Client 11.0};SERVER="
                      + settings.server +
                      ";DATABASE=" +
                      settings.database +
                      ";Trusted_Connection=yes;")


# start execution timer
start = time.time()

# get MGRA list for ABM scenario
mgra_list_sql = 'SELECT [user_id], [id], [value] from [popsyn_3_0].[sbreport].[mgra] where [user_id] = system_user order by id'

mgra_result = pd.read_sql_query(
    sql=mgra_list_sql,
    con=conn
)

print("Beginning [scenario_id]: " + scenario_id)
scenario_id = int(scenario_id)

print("Running: Report queries")

# for each sheet of report queries --------------------
for qry in settings.report_queries:

    print("Running: " + qry["sp"] + " " + qry["sheet"])

    # execute stored procedure with specified arguments
    result = pd.read_sql_query(
        sql="EXECUTE " + qry["sp"] + " " + qry["args"],
        con=conn,
        params=[scenario_id]
    )

    # validate scenario id (i.e. not hard-coded in stored procedure)
    for s in result["scenario_id"]:
        if s != scenario_id:
            msg = ("Invalid scenario id:",s,"should be:", scenario_id, "for sheet:", qry["sheet"])
            print(result)
            raise ValueError(msg)

    # set first row of sql query to be column headers
    new_header = result.iloc[0]  # grab the first row for the header
    result = result[1:]  # take the data less the header row
    result.columns = new_header  # set the header row as the df header

    # set columns to numeric where possible
    result = result.apply(pd.to_numeric) # convert columns to numeric

    # create blank worksheet (delete rows from template that have data)
    worksheet = templateWriter.sheets[qry["sheet"]]
    worksheet.delete_rows(2, worksheet.max_row-1)

    # write results to specified Excel template sheet
    result.to_excel(
        excel_writer=templateWriter,
        sheet_name=qry["sheet"],
        na_rep="NULL",
        header=True,
        index=False,
        startrow=1,
        engine="openpyxl")

    # align cells for pretty output
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')


print("Elapsed time: %5.2f minutes" % ((time.time() - start) / 60.0))

# if protection option set
if settings.template_protect is True:
    # for each worksheet in the template
    for sheet in template.worksheets:
        # set protection and password
        sheet.protection.set_password(settings.template_password)

# save the completed template
template.save(settings.report_write_path + " - " + file_suffix + ".xlsx")
